#%%
from arcgis.gis import GIS
from datetime import datetime
import pandas as pd
import arcpy
#%%ontract = arcpy.GetParameterAsText(0)
contract = '10-100'
#%%

portal_url = 'https://arcportal-ucop-corps.usace.army.mil/s0portal'
un = ""
pw = ""
gis = GIS(portal_url, un, pw, verify_cert=False)
token = gis._con._token

#%%
layer_item = gis.content.get('bb84179b14514c44a2cb2f9fe133f7c0')
layers = layer_item.layers
tables = layer_item.tables
sale_flayer = layers[0]
sale_feature = sale_flayer.query(where = f"contractid = '{contract}'").features[0]
bids_flayer = tables[3]
bids_df = bids_flayer.query(where = f"parentglobalid = '{sale_feature.attributes['globalid']}'").sdf


#%%
from openpyxl import load_workbook
xls = r"C:\Users\g3retjjj\Downloads\Abstract of Bids.xlsx"
wb = load_workbook(xls)
ws = wb['Sheet1']
#ws['A1'] = 'A1'
#wb.save(f"{ContractTitle} Abstract of Bids.xlsx")
#%%
sale_feature.attributes['globalid']
# %%
bids_df
# %%
ws['C5']='TEST LALALALA'
wb.save(r"C:\Users\g3retjjj\Downloads\Abstract of BidsTEST.xlsx")
# %%
